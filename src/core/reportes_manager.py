from src.db.database import get_supabase
from src.core.auth_manager import AuthManager
from collections import defaultdict

class ReportesManager:
    @staticmethod
    def _check_permission():
        if not AuthManager.is_admin():
            raise PermissionError("Solo la dueña puede acceder a los reportes de ganancias y estadísticas.")

    @staticmethod
    def get_vendedores():
        ReportesManager._check_permission()
        supabase = get_supabase()
        res = supabase.table('usuarios').select('id, nombre, username').eq('activo', True).execute()
        return res.data

    @staticmethod
    def get_ventas_por_fecha(fecha_desde: str, fecha_hasta: str, metodo_pago: str = None, usuario_id: str = None):
        ReportesManager._check_permission()
        desde = f"{fecha_desde}T00:00:00"
        hasta = f"{fecha_hasta}T23:59:59"
        
        supabase = get_supabase()
        
        query = supabase.table('ventas').select('id, fecha, total, metodo_pago, clientes(nombre), usuarios(nombre, username)').neq('estado', 'CANCELADA').gte('fecha', desde).lte('fecha', hasta)
        
        if metodo_pago:
            query = query.eq('metodo_pago', metodo_pago)
            
        if usuario_id:
            query = query.eq('usuario_id', usuario_id)
            
        res = query.order('fecha', desc=True).execute()
        ventas = res.data
        
        efectivo = 0.0
        transferencia = 0.0
        
        ventas_fmt = []
        for v in ventas:
            total = float(v['total'])
            mp = v['metodo_pago']
            
            if mp == 'EFECTIVO':
                efectivo += total
            elif mp in ('TRANSFERENCIA', 'TARJETA', 'TARJETA/TRANSFERENCIA'):
                transferencia += total
            elif mp == 'MIXTO':
                # En Supabase Python habría que traer los movimientos asociados para precisión,
                # por simplicidad asumiremos todo efectivo aquí o consultar caja_movimientos.
                # Para evitar N+1 queries, simplificaremos
                pass
            
            vendedor = 'Sistema'
            if v.get('usuarios'):
                vendedor = v['usuarios'].get('nombre') or v['usuarios'].get('username') or 'Desconocido'
                
            ventas_fmt.append({
                'id': v['id'],
                'fecha': v['fecha'],
                'total': total,
                'metodo_pago': mp,
                'cliente': v['clientes']['nombre'] if v.get('clientes') else 'Consumidor Final',
                'vendedor': vendedor
            })
            
        return {
            'ventas': ventas_fmt,
            'total_efectivo': efectivo,
            'total_transferencia': transferencia,
            'total_general': sum(v['total'] for v in ventas_fmt)
        }

    @staticmethod
    def get_productos_mas_vendidos(fecha_desde: str, fecha_hasta: str):
        ReportesManager._check_permission()
        desde = f"{fecha_desde}T00:00:00"
        hasta = f"{fecha_hasta}T23:59:59"
        
        supabase = get_supabase()
        res = supabase.table('ventas_detalle').select(
            'cantidad, subtotal, productos(codigo_barras, nombre), ventas!inner(estado, fecha)'
        ).neq('ventas.estado', 'CANCELADA').gte('ventas.fecha', desde).lte('ventas.fecha', hasta).execute()
        
        agrupado = defaultdict(lambda: {'codigo_barras': '', 'nombre': '', 'cant_total': 0, 'recaudacion': 0.0})
        
        for d in res.data:
            if not d.get('productos'):
                continue
            prod_nombre = d['productos']['nombre']
            prod_codigo = d['productos']['codigo_barras']
            
            agrupado[prod_nombre]['codigo_barras'] = prod_codigo
            agrupado[prod_nombre]['nombre'] = prod_nombre
            agrupado[prod_nombre]['cant_total'] += float(d['cantidad'])
            agrupado[prod_nombre]['recaudacion'] += float(d['subtotal'])
            
        resultado = list(agrupado.values())
        resultado.sort(key=lambda x: x['cant_total'], reverse=True)
        return resultado

    @staticmethod
    def get_cierres_caja(fecha_desde: str, fecha_hasta: str, usuario_id: str = None):
        ReportesManager._check_permission()
        desde = f"{fecha_desde}T00:00:00"
        hasta = f"{fecha_hasta}T23:59:59"
        
        supabase = get_supabase()
        try:
            query = supabase.table('caja_sesiones').select('*, usuarios(nombre, username)').gte('fecha_apertura', desde).lte('fecha_apertura', hasta)
            if usuario_id:
                query = query.eq('usuario_id', usuario_id)
            res = query.order('fecha_apertura', desc=True).execute()
            return res.data
        except Exception:
            query = supabase.table('caja_sesiones').select('*').gte('fecha_apertura', desde).lte('fecha_apertura', hasta)
            res = query.order('fecha_apertura', desc=True).execute()
            return res.data

    @staticmethod
    def get_alertas_reposicion():
        ReportesManager._check_permission()
        supabase = get_supabase()
        res = supabase.table('productos').select('codigo_barras, nombre, stock_actual, stock_minimo, stock_maximo').eq('activo', True).execute()
        
        alertas = []
        for p in res.data:
            actual = float(p['stock_actual'])
            minimo = float(p['stock_minimo'])
            maximo = float(p['stock_maximo'])
            
            if actual <= minimo:
                alertas.append({
                    'codigo_barras': p['codigo_barras'],
                    'nombre': p['nombre'],
                    'stock_actual': actual,
                    'stock_minimo': minimo,
                    'stock_maximo': maximo,
                    'sugerido_pedir': maximo - actual
                })
                
        alertas.sort(key=lambda x: x['sugerido_pedir'], reverse=True)
        return alertas

    @staticmethod
    def get_reporte_ganancias(fecha_desde: str, fecha_hasta: str, usuario_id: str = None):
        ReportesManager._check_permission()
        desde = f"{fecha_desde}T00:00:00"
        hasta = f"{fecha_hasta}T23:59:59"
        
        supabase = get_supabase()
        query = supabase.table('ventas_detalle').select(
            'cantidad, costo_unitario, subtotal, ventas!inner(estado, fecha, usuario_id)'
        ).neq('ventas.estado', 'CANCELADA').gte('ventas.fecha', desde).lte('ventas.fecha', hasta)
        
        if usuario_id:
            query = query.eq('ventas.usuario_id', usuario_id)
            
        res = query.execute()
        
        ganancias_por_dia = defaultdict(lambda: {'total_vendido': 0.0, 'costo_total': 0.0, 'ganancia_neta': 0.0})
        totales = {'total_vendido': 0.0, 'costo_total': 0.0, 'ganancia_neta': 0.0}
        
        for d in res.data:
            fecha_str = d['ventas']['fecha']
            dia = fecha_str[:10]
                
            subt = float(d['subtotal'])
            costo = float(d['costo_unitario']) * float(d['cantidad'])
            neta = subt - costo
            
            ganancias_por_dia[dia]['total_vendido'] += subt
            ganancias_por_dia[dia]['costo_total'] += costo
            ganancias_por_dia[dia]['ganancia_neta'] += neta
            
            totales['total_vendido'] += subt
            totales['costo_total'] += costo
            totales['ganancia_neta'] += neta
            
        resultado_dias = []
        for dia, vals in ganancias_por_dia.items():
            resultado_dias.append({
                'dia': dia,
                'total_vendido': vals['total_vendido'],
                'costo_total': vals['costo_total'],
                'ganancia_neta': vals['ganancia_neta']
            })
            
        resultado_dias.sort(key=lambda x: x['dia'], reverse=True)
        
        return {
            'ganancias_por_dia': resultado_dias,
            'totales': totales
        }

    @staticmethod
    def get_ventas_por_rubro(fecha_desde: str, fecha_hasta: str):
        ReportesManager._check_permission()
        desde = f"{fecha_desde}T00:00:00"
        hasta = f"{fecha_hasta}T23:59:59"
        
        supabase = get_supabase()
        res = supabase.table('ventas_detalle').select(
            'cantidad, costo_unitario, subtotal, productos(categorias(nombre)), ventas!inner(estado, fecha)'
        ).neq('ventas.estado', 'CANCELADA').gte('ventas.fecha', desde).lte('ventas.fecha', hasta).execute()
        
        agrupado = defaultdict(lambda: {'total_vendido': 0.0, 'costo_total': 0.0, 'ganancia_neta': 0.0})
        
        for d in res.data:
            rubro = 'Sin Rubro'
            if d.get('productos') and d['productos'].get('categorias'):
                rubro = d['productos']['categorias']['nombre'] or 'Sin Rubro'
                
            subt = float(d['subtotal'])
            costo = float(d['costo_unitario']) * float(d['cantidad'])
            neta = subt - costo
            
            agrupado[rubro]['total_vendido'] += subt
            agrupado[rubro]['costo_total'] += costo
            agrupado[rubro]['ganancia_neta'] += neta
            
        resultado = []
        for r, vals in agrupado.items():
            resultado.append({
                'rubro': r,
                'total_vendido': vals['total_vendido'],
                'costo_total': vals['costo_total'],
                'ganancia_neta': vals['ganancia_neta']
            })
            
        resultado.sort(key=lambda x: x['total_vendido'], reverse=True)
        return resultado

    @staticmethod
    def generar_excel_declaracion_ventas(fecha_desde: str, fecha_hasta: str, filepath: str):
        ReportesManager._check_permission()
        desde = f"{fecha_desde}T00:00:00"
        hasta = f"{fecha_hasta}T23:59:59"
        
        supabase = get_supabase()
        res = supabase.table('ventas').select(
            'id, fecha, total, metodo_pago, clientes(nombre), usuarios(nombre, username)'
        ).neq('estado', 'CANCELADA').gte('fecha', desde).lte('fecha', hasta).order('fecha', desc=False).execute()
        
        ventas = res.data
        
        # Agrupar por método de pago
        grupos = defaultdict(list)
        totales_grupo = defaultdict(float)
        
        for v in ventas:
            mp = v.get('metodo_pago') or 'OTROS'
            total = float(v['total'])
            vendedor = 'Sistema'
            if v.get('usuarios'):
                vendedor = v['usuarios'].get('nombre') or v['usuarios'].get('username') or 'Sistema'
            cliente = v['clientes']['nombre'] if v.get('clientes') else 'Consumidor Final'
            
            # Formatear fecha limpia
            fecha_str = v['fecha']
            if 'T' in fecha_str:
                partes = fecha_str.split('T')
                f_date = partes[0]
                f_time = partes[1].split('.')[0]
                fecha_str = f"{f_date} {f_time}"

            item = {
                'id': f"{v['id']:08d}",
                'fecha': fecha_str,
                'cliente': cliente,
                'vendedor': vendedor,
                'total': total
            }
            grupos[mp].append(item)
            totales_grupo[mp] += total

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            # Hoja Resumen
            ws_resumen = wb.active
            ws_resumen.title = "Resumen de Declaración"
            
            # Estilos
            font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
            font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_bold = Font(name="Calibri", size=11, bold=True)
            
            fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            fill_subtotal = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )

            # Título Resumen
            ws_resumen['A1'] = "DECLARACIÓN DE VENTAS POR MEDIO DE PAGO"
            ws_resumen['A1'].font = font_title
            ws_resumen['A2'] = f"Período: {fecha_desde} al {fecha_hasta}"
            ws_resumen['A2'].font = font_subtitle
            
            ws_resumen.append([])
            headers_resumen = ["Medio de Pago", "Cantidad de Ventas", "Total Recaudado ($)"]
            ws_resumen.append(headers_resumen)
            
            for col in range(1, 4):
                cell = ws_resumen.cell(row=4, column=col)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row_idx = 5
            grand_total = 0.0
            grand_cant = 0
            
            for mp, items in grupos.items():
                tot = totales_grupo[mp]
                cant = len(items)
                grand_total += tot
                grand_cant += cant
                
                c1 = ws_resumen.cell(row=row_idx, column=1, value=mp)
                c2 = ws_resumen.cell(row=row_idx, column=2, value=cant)
                c3 = ws_resumen.cell(row=row_idx, column=3, value=tot)
                
                c3.number_format = '"$"#,##0.00'
                for c in (c1, c2, c3):
                    c.border = thin_border
                row_idx += 1
                
            # Fila Total General
            c1 = ws_resumen.cell(row=row_idx, column=1, value="TOTAL GENERAL")
            c2 = ws_resumen.cell(row=row_idx, column=2, value=grand_cant)
            c3 = ws_resumen.cell(row=row_idx, column=3, value=grand_total)
            c3.number_format = '"$"#,##0.00'
            
            for c in (c1, c2, c3):
                c.font = font_bold
                c.fill = fill_subtotal
                c.border = thin_border
                
            ws_resumen.column_dimensions['A'].width = 25
            ws_resumen.column_dimensions['B'].width = 20
            ws_resumen.column_dimensions['C'].width = 22

            # Crear una hoja por cada método de pago
            for mp, items in grupos.items():
                sheet_name = mp.replace('/', '-').replace('\\', '-')[:30]
                ws = wb.create_sheet(title=sheet_name)
                
                ws['A1'] = f"VENTAS - {mp}"
                ws['A1'].font = font_title
                ws['A2'] = f"Período: {fecha_desde} al {fecha_hasta}"
                ws['A2'].font = font_subtitle
                ws.append([])
                
                headers = ["Factura Nº", "Fecha y Hora", "Cliente", "Vendedor", "Total ($)"]
                ws.append(headers)
                
                for col in range(1, 6):
                    cell = ws.cell(row=4, column=col)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                r = 5
                subtot = 0.0
                for it in items:
                    c1 = ws.cell(row=r, column=1, value=it['id'])
                    c2 = ws.cell(row=r, column=2, value=it['fecha'])
                    c3 = ws.cell(row=r, column=3, value=it['cliente'])
                    c4 = ws.cell(row=r, column=4, value=it['vendedor'])
                    c5 = ws.cell(row=r, column=5, value=it['total'])
                    c5.number_format = '"$"#,##0.00'
                    
                    for c in (c1, c2, c3, c4, c5):
                        c.border = thin_border
                    subtot += it['total']
                    r += 1
                    
                # Total por sección
                ws.cell(row=r, column=1, value="TOTAL").font = font_bold
                c_sub = ws.cell(row=r, column=5, value=subtot)
                c_sub.font = font_bold
                c_sub.number_format = '"$"#,##0.00'
                ws.cell(row=r, column=1).fill = fill_subtotal
                c_sub.fill = fill_subtotal
                
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 22
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 18

            wb.save(filepath)
            return True
            
        except ImportError:
            # Fallback a CSV tabulado con UTF-8 BOM para Excel si openpyxl no estuviera instalado
            import csv
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(["DECLARACIÓN DE VENTAS POR MEDIO DE PAGO"])
                writer.writerow([f"Período: {fecha_desde} al {fecha_hasta}"])
                writer.writerow([])
                
                for mp, items in grupos.items():
                    writer.writerow([f"=== SECCIÓN: {mp} ==="])
                    writer.writerow(["Factura Nro", "Fecha y Hora", "Cliente", "Vendedor", "Total"])
                    subtot = 0.0
                    for it in items:
                        writer.writerow([it['id'], it['fecha'], it['cliente'], it['vendedor'], f"{it['total']:.2f}"])
                        subtot += it['total']
                    writer.writerow(["TOTAL SECCIÓN", "", "", "", f"{subtot:.2f}"])
                    writer.writerow([])
            return True

